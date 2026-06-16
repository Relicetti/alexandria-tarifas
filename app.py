from dotenv import load_dotenv
import os
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"), override=True)

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import io
import json
import threading
import uuid
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
import db
from calculos import calcular, CAMPOS_GRUPO
from concessionarias import CONCESSIONARIAS, normalizar_distribuidora

app = Flask(__name__)
app.secret_key = "alexandria-tarifas-2024"

db.init_db()


_DATA_DIR     = Path(db.DB_PATH).parent          # volume /data no Railway, pasta local caso contrário
FEEDBACK_FILE = _DATA_DIR / "feedback_extracao.jsonl"
DEBUG_LOG     = _DATA_DIR / "debug_feedback.log"
_EXTR_DIR     = _DATA_DIR / "extraidos"
_EXTR_DIR.mkdir(exist_ok=True)


def _salvar_extraido(dados: dict) -> str:
    token = uuid.uuid4().hex
    (_EXTR_DIR / f"{token}.json").write_text(
        json.dumps(dados, ensure_ascii=False), encoding="utf-8"
    )
    return token


def _carregar_extraido(token: str) -> dict | None:
    p = _EXTR_DIR / f"{token}.json"
    if not p.exists():
        return None
    dados = json.loads(p.read_text(encoding="utf-8"))
    p.unlink()   # usa uma vez só
    return dados


def _log_debug(msg: str):
    with open(DEBUG_LOG, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")

# Campos que a IA extrai e o usuário pode corrigir
_CAMPOS_EXTRACAO = [
    "consumo_kwh", "injetada_kwh", "valor_concessionaria",
    "te_consumo", "tusd_consumo", "te_compensada", "tusd_compensada",
    "tarifa_distribuidora_input", "tarifa_compensada_input", "ajuste_gd2",
    "tusd_distribuidora", "te_distribuidora", "desconto_injecao",
    "scee_consumo", "scee_injecao", "scee_comp_nao_isento",
    "scee_beneficio_bruto", "scee_beneficio_liquido",
    "tusd_injetada_gd", "te_injetada_gd", "tusd_fornecida_gd", "te_fornecida_gd",
    "b_amarela_cons_valor", "b_verm_p1_cons_valor", "b_verm_p2_cons_valor",
    "b_amarela_inj_valor", "b_verm_p1_inj_valor", "b_verm_p2_inj_valor",
    "b_amarela_cons_kwh", "b_verm_p1_cons_kwh", "b_verm_p2_cons_kwh",
    "aliquota_icms", "valor_icms", "aliquota_pis", "valor_pis",
    "aliquota_cofins", "valor_cofins",
    "grupo",
]


def _registrar_feedback(extraido_orig: dict, dados_salvos: dict):
    """Compara extração original vs valores salvos e registra divergências."""
    diffs = {}
    for campo in _CAMPOS_EXTRACAO:
        v_ext = extraido_orig.get(campo)
        v_cor = dados_salvos.get(campo)

        # Normaliza para float quando possível
        def _n(v):
            if v in (None, "", 0, 0.0):
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return v

        ve, vc = _n(v_ext), _n(v_cor)
        if ve != vc:
            diffs[campo] = {"extraido": ve, "corrigido": vc}

    registro = {
        "ts": datetime.now().isoformat(),
        "distribuidora": extraido_orig.get("distribuidora", ""),
        "grupo": extraido_orig.get("grupo", ""),
        "diffs": diffs,
    }

    with open(FEEDBACK_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(registro, ensure_ascii=False) + "\n")

    if diffs:
        print(f"[feedback] {len(diffs)} campo(s) corrigido(s) — melhorando prompt em background...")
        def _melhorar():
            try:
                from melhorar_prompt import melhorar
                melhorar()
            except Exception as e:
                print(f"[feedback] Erro ao melhorar prompt: {e}")
        threading.Thread(target=_melhorar, daemon=True).start()
    else:
        print("[feedback] Extração sem divergências — nenhuma correção necessária.")


def mes_atual():
    hoje = date.today()
    return f"{hoje.year}-{hoje.month:02d}-01"


def _float(v, default=None):
    try:
        return float(v) if v not in (None, '') else default
    except (ValueError, TypeError):
        return default


def _parse_form(form):
    """Extrai e converte todos os campos do formulário."""
    f = form
    cobra = 1 if f.get("cobra_band") else 0
    d = {
        "cliente_id":               None,
        "usina_id":                 f.get("usina_id", ""),
        "distribuidora":            f.get("distribuidora", ""),
        "instalacao":               f.get("instalacao", ""),
        "grupo":                    f.get("grupo", ""),
        "tipo_gd":                  f.get("tipo_gd", "GD1"),
        "modalidade":               f.get("modalidade", "Geração Compartilhada"),
        "mes_referencia":           f["mes_referencia"][:7] + "-01",
        "valor_concessionaria":     _float(f.get("valor_concessionaria")),
        "consumo_kwh":              _float(f.get("consumo_kwh")),
        "injetada_kwh":             _float(f.get("injetada_kwh")),
        "desconto_base":            _float(f.get("desconto_base_pct"), 0) / 100,
        "desconto_aplicado":        _float(f.get("desconto_aplicado_pct"), 0) / 100 or None,
        "cobra_band":               cobra,
        # GER
        "te_consumo":               _float(f.get("te_consumo")),
        "tusd_consumo":             _float(f.get("tusd_consumo")),
        "te_compensada":            _float(f.get("te_compensada")),
        "tusd_compensada":          _float(f.get("tusd_compensada")),
        # NEOENERGIA
        "tusd_distribuidora":       _float(f.get("tusd_distribuidora")),
        "te_distribuidora":         _float(f.get("te_distribuidora")),
        "desconto_injecao":         _float(f.get("desconto_injecao")),
        # Direto (EQT/ENERGISA/LIGHT/CEMIG/BRASILIA)
        "tarifa_distribuidora_input": _float(f.get("tarifa_distribuidora_input")),
        "tarifa_compensada_input":  _float(f.get("tarifa_compensada_input")),
        "ajuste_gd2":               _float(f.get("ajuste_gd2"), 0),
        # EQT SCEE
        "scee_consumo":             _float(f.get("scee_consumo")),
        "scee_injecao":             _float(f.get("scee_injecao")),
        "scee_comp_nao_isento":     _float(f.get("scee_comp_nao_isento")),
        "scee_beneficio_bruto":     _float(f.get("scee_beneficio_bruto"), 0),
        "scee_beneficio_liquido":   _float(f.get("scee_beneficio_liquido"), 0),
        # LIGHT
        "tusd_injetada_gd":         _float(f.get("tusd_injetada_gd")),
        "te_injetada_gd":           _float(f.get("te_injetada_gd")),
        "tusd_fornecida_gd":        _float(f.get("tusd_fornecida_gd")),
        "te_fornecida_gd":          _float(f.get("te_fornecida_gd")),
        "aliquota_icms":            _float(f.get("aliquota_icms"), 0),
        "valor_icms":               _float(f.get("valor_icms"), 0),
        "aliquota_pis":             _float(f.get("aliquota_pis"), 0),
        "valor_pis":                _float(f.get("valor_pis"), 0),
        "aliquota_cofins":          _float(f.get("aliquota_cofins"), 0),
        "valor_cofins":             _float(f.get("valor_cofins"), 0),
        # Bandeira consumo
        "b_amarela_cons_kwh":       _float(f.get("b_amarela_cons_kwh"), 0),
        "b_amarela_cons_valor":     _float(f.get("b_amarela_cons_valor"), 0),
        "b_verm_p1_cons_kwh":       _float(f.get("b_verm_p1_cons_kwh"), 0),
        "b_verm_p1_cons_valor":     _float(f.get("b_verm_p1_cons_valor"), 0),
        "b_verm_p2_cons_kwh":       _float(f.get("b_verm_p2_cons_kwh"), 0),
        "b_verm_p2_cons_valor":     _float(f.get("b_verm_p2_cons_valor"), 0),
        # Bandeira injeção
        "b_amarela_inj_kwh":        _float(f.get("b_amarela_inj_kwh"), 0),
        "b_amarela_inj_valor":      _float(f.get("b_amarela_inj_valor"), 0),
        "b_verm_p1_inj_kwh":        _float(f.get("b_verm_p1_inj_kwh"), 0),
        "b_verm_p1_inj_valor":      _float(f.get("b_verm_p1_inj_valor"), 0),
        "b_verm_p2_inj_kwh":        _float(f.get("b_verm_p2_inj_kwh"), 0),
        "b_verm_p2_inj_valor":      _float(f.get("b_verm_p2_inj_valor"), 0),
        "status_pagamento":         "pendente",
        "data_pagamento":           None,
        "obs":                      "",
    }
    return d


# ── DASHBOARD ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    mes          = request.args.get("mes", mes_atual())
    grupo        = request.args.get("grupo", "")
    distribuidora= request.args.get("distribuidora", "")

    faturas = db.get_faturas_mes(mes) if mes else db.get_todas_faturas()
    if grupo:
        faturas = [f for f in faturas if f["grupo"] == grupo]
    if distribuidora:
        faturas = [f for f in faturas if f["distribuidora"] == distribuidora]

    meses = db.get_meses_disponiveis()
    if mes and mes not in meses:
        meses = [mes] + meses

    return render_template("index.html", faturas=faturas, mes=mes, meses=meses,
                           grupo=grupo, grupos=db.GRUPOS,
                           distribuidora=distribuidora,
                           distribuidoras=db.get_distribuidoras())


# ── FATURAS ───────────────────────────────────────────────────────────────────

@app.route("/faturas/nova", methods=["GET", "POST"])
@app.route("/faturas/<int:id>/editar", methods=["GET", "POST"])
def form_fatura(id=None):
    fatura = db.get_fatura(id) if id else None

    if request.method == "POST":
        data  = _parse_form(request.form)
        grupo = data["grupo"]

        from concessionarias import normalizar_distribuidora
        data["distribuidora"] = normalizar_distribuidora(data["distribuidora"])

        extraido_json = request.form.get("_extraido_json")
        _log_debug(f"_extraido_json presente={bool(extraido_json)} tamanho={len(extraido_json) if extraido_json else 0}")
        if extraido_json:
            try:
                extraido_orig = json.loads(extraido_json)
                extraido_orig["grupo"] = extraido_orig.get("grupo") or grupo
                _registrar_feedback(extraido_orig, data)
            except Exception as e:
                import traceback
                _log_debug(f"ERRO feedback: {e}\n{traceback.format_exc()}")

        try:
            resultado = calcular(grupo, data)
            data.update(resultado)
            if not id:
                existente = db.get_fatura_por_instalacao_mes(
                    data["instalacao"], data["mes_referencia"]
                )
                if existente:
                    db.salvar_fatura(data, existente["id"])
                    flash("Fatura existente atualizada.", "success")
                    return redirect(url_for("index", mes=data["mes_referencia"]))
            db.salvar_fatura(data, id)
            flash("Fatura salva.", "success")
            return redirect(url_for("index", mes=data["mes_referencia"]))
        except Exception as e:
            import traceback
            _log_debug(f"ERRO SALVAR: {e}\n{traceback.format_exc()}")
            flash(f"Erro no cálculo: {e}", "danger")

    token    = request.args.get("_extr")
    extraido = _carregar_extraido(token) if token else None
    _log_debug(f"GET form_fatura: token={token} extraido={'SIM keys='+str(list(extraido.keys())) if extraido else 'NAO'}")
    mes_pre   = request.args.get("mes", mes_atual())
    grupo_pre = None

    if extraido:
        mes_pre   = (extraido.get("mes_referencia") or mes_pre[:7]) + "-01"
        grupo_pre = extraido.get("_grupo") or extraido.get("grupo")
    elif fatura:
        grupo_pre = fatura["grupo"]

    return render_template("form_fatura.html",
                           fatura=fatura,
                           mes_pre=mes_pre,
                           grupo_pre=grupo_pre,
                           campos_grupo=CAMPOS_GRUPO,
                           grupos=db.GRUPOS,
                           concessionarias=CONCESSIONARIAS,
                           extraido=extraido)


@app.route("/faturas/<int:id>/deletar", methods=["POST"])
def deletar_fatura(id):
    fatura = db.get_fatura(id)
    mes = fatura["mes_referencia"] if fatura else mes_atual()
    db.deletar_fatura(id)
    flash("Fatura removida.", "warning")
    return redirect(url_for("index", mes=mes))


# ── UPLOAD + EXTRAÇÃO ────────────────────────────────────────────────────────

@app.route("/faturas/upload", methods=["GET", "POST"])
def upload_fatura():
    if request.method == "POST":
        arq = request.files.get("fatura_pdf")
        if not arq or not arq.filename.lower().endswith(".pdf"):
            flash("Selecione um arquivo PDF.", "danger")
            return render_template("upload_fatura.html")
        try:
            from extrator import extrair_fatura
            dados = extrair_fatura(arq.read())
            # normaliza nome da distribuidora para o nome curto da lista
            dados["distribuidora"] = normalizar_distribuidora(dados.get("distribuidora", ""))
            instalacao = str(dados.get("instalacao", "")).strip()
            if instalacao:
                f_existente = db.get_faturas_instalacao(instalacao)
                if f_existente:
                    dados["_grupo"] = dados.get("_grupo") or f_existente[0].get("grupo")
            if not dados.get("_grupo") and dados.get("grupo"):
                dados["_grupo"] = dados["grupo"]
            token = _salvar_extraido(dados)
            _log_debug(f"UPLOAD OK: dist={dados.get('distribuidora')} inst={dados.get('instalacao')} chaves={list(dados.keys())} token={token}")
            flash("✅ Dados extraídos! Revise e salve.", "success")
            return redirect(url_for("form_fatura", _extr=token))
        except ValueError as e:
            flash(str(e), "danger")
        except Exception as e:
            flash(f"Erro na extração: {e}", "danger")
    return render_template("upload_fatura.html")


# ── HISTÓRICO ────────────────────────────────────────────────────────────────

@app.route("/historico/<instalacao>")
def historico(instalacao):
    faturas = db.get_faturas_instalacao(instalacao)
    return render_template("historico.html", instalacao=instalacao, faturas=faturas)


# ── EXPORTAR EXCEL ───────────────────────────────────────────────────────────

@app.route("/exportar")
def exportar():
    mes          = request.args.get("mes", "")
    grupo        = request.args.get("grupo", "")
    distribuidora= request.args.get("distribuidora", "")

    faturas = db.get_faturas_mes(mes) if mes else db.get_todas_faturas()
    if grupo:
        faturas = [f for f in faturas if f["grupo"] == grupo]
    if distribuidora:
        faturas = [f for f in faturas if f["distribuidora"] == distribuidora]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faturas"

    headers = ["Grupo", "Distribuidora", "Instalação", "Usina", "Mês Ref.",
               "Tipo GD", "Modalidade",
               "Consumo (kWh)", "Injetado (kWh)", "Tarifa Dist. (R$/kWh)",
               "Tarifa Comp. (R$/kWh)", "Tarifa Geração (R$/kWh)",
               "Desconto Real (%)", "Desconto Ref. (%)", "Valor Conc. (R$)"]

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for col_i, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for f in faturas:
        ws.append([
            f["grupo"],
            f["distribuidora"],
            f["instalacao"],
            f["usina_id"],
            f["mes_referencia"][:7],
            f["tipo_gd"]   if "tipo_gd"   in f.keys() else "",
            f["modalidade"] if "modalidade" in f.keys() else "",
            f["consumo_kwh"],
            f["injetada_kwh"],
            round(f["tarifa_distribuidora"], 6) if f["tarifa_distribuidora"] else None,
            round(f["tarifa_compensada"], 6)    if f["tarifa_compensada"]    else None,
            round(f["tarifa_geracao"], 6)        if f["tarifa_geracao"]       else None,
            round(f["desconto_real"] * 100, 2)  if f["desconto_real"] is not None else None,
            round(f["desconto_ref"]  * 100, 2)  if f["desconto_ref"]         else None,
            f["valor_concessionaria"],
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    partes = ["faturas", mes[:7] if mes else "todos"]
    if grupo:        partes.append(grupo)
    if distribuidora:
        partes.append(distribuidora.replace(" ", "_"))
    nome = "_".join(partes) + ".xlsx"
    return send_file(output, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── TARIFA GERADOR ───────────────────────────────────────────────────────────

@app.route("/tarifa-gerador")
def tarifa_gerador():
    db.init_tarifas_gerador()
    distribuidoras = db.get_distribuidoras()
    meses          = db.get_meses_disponiveis()
    salvos         = db.get_tarifas_gerador()
    return render_template("tarifa_gerador.html",
                           distribuidoras=distribuidoras,
                           meses=meses,
                           salvos=salvos)


@app.route("/api/tarifa-gerador/salvar", methods=["POST"])
def api_salvar_tarifa_gerador():
    from flask import jsonify
    db.init_tarifas_gerador()
    data = request.get_json()
    if not data:
        return jsonify({"erro": "Dados inválidos"}), 400
    try:
        db.salvar_tarifa_gerador(data)
        # Buscar criado_em do registro salvo
        rows = db.get_tarifas_gerador(distribuidora=data.get("distribuidora"), mes=data.get("mes_referencia"))
        criado_em = None
        for r in rows:
            if r["tipo_gd"] == data.get("tipo_gd") and r["modalidade"] == data.get("modalidade"):
                criado_em = r["criado_em"]
                break
        return jsonify({"ok": True, "criado_em": criado_em})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/tarifa-gerador/<int:id>/deletar", methods=["POST"])
def api_deletar_tarifa_gerador(id):
    from flask import jsonify
    db.deletar_tarifa_gerador(id)
    return jsonify({"ok": True})


@app.route("/tarifa-gerador/exportar")
def exportar_tarifa_gerador():
    db.init_tarifas_gerador()
    salvos = db.get_tarifas_gerador()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarifas Gerador"

    headers = ["Distribuidora", "Mês", "Tipo GD", "Modalidade",
               "Desconto GD (%)", "Tar. Compensada (R$/kWh)",
               "Tar. Distribuidora (R$/kWh)", "Deságio (%)", "T_gerador (R$/kWh)",
               "Salvo em"]

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for col_i, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    modal_label = {"AC": "Autoconsumo", "GC": "Geração Compartilhada"}
    for s in salvos:
        ws.append([
            s["distribuidora"],
            s["mes_referencia"][:7],
            s["tipo_gd"] or "",
            modal_label.get(s["modalidade"], s["modalidade"] or ""),
            round(s["desconto_gd"] * 100, 4)       if s["desconto_gd"]          else None,
            round(s["tarifa_compensada"],   6)      if s["tarifa_compensada"]    else None,
            round(s["tarifa_distribuidora"],6)      if s["tarifa_distribuidora"] else None,
            round(s["desagio"] * 100, 4)            if s["desagio"]              else None,
            round(s["t_gerador"],           6)      if s["t_gerador"]            else None,
            s["criado_em"][:16] if s["criado_em"] else "",
        ])

    # Formata colunas numéricas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"
            if cell.column in (5, 8):          # % colunas
                cell.number_format = "0.00%"
                if cell.value is not None:
                    cell.value = cell.value / 100  # transforma de volta para decimal

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name="tarifas_gerador.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/tarifa-gerador")
def api_tarifa_gerador():
    from flask import jsonify
    dist    = request.args.get("distribuidora", "").strip()
    mes     = request.args.get("mes", "").strip()       # YYYY-MM
    tipo_gd = request.args.get("tipo_gd", "").strip()   # GD1 / GD2 / ""
    modal   = request.args.get("modalidade", "").strip() # AC / GC / ""

    if not dist or not mes:
        return jsonify({"erro": "Selecione distribuidora e mês"}), 400

    # Mês pode chegar como YYYY-MM; no banco está YYYY-MM-01
    mes_db = mes[:7] + "-01"

    cond   = ["f.distribuidora = ?", "f.mes_referencia = ?"]
    params = [dist, mes_db]

    if tipo_gd:
        cond.append("f.tipo_gd = ?")
        params.append(tipo_gd)

    if modal == "AC":
        cond.append("f.modalidade = 'Autoconsumo'")
    elif modal == "GC":
        cond.append("f.modalidade = 'Geração Compartilhada'")

    where = " AND ".join(cond)

    with db.get_conn() as conn:
        row = conn.execute(f"""
            SELECT
                AVG(f.tarifa_compensada)    AS tarifa_compensada,
                AVG(f.tarifa_distribuidora) AS tarifa_distribuidora,
                AVG(f.tarifa_geracao)       AS tarifa_geracao,
                COUNT(*)                    AS total,
                COUNT(CASE WHEN f.instalacao NOT LIKE 'HIST-%' THEN 1 END) AS total_reais
            FROM faturas f
            WHERE {where}
              AND f.tarifa_compensada    IS NOT NULL
              AND f.tarifa_distribuidora IS NOT NULL
        """, params).fetchone()

        # Desconto GD: média das faturas reais da distribuidora
        cond_desc   = ["distribuidora = ?", "instalacao NOT LIKE 'HIST-%'", "desconto_base IS NOT NULL"]
        params_desc = [dist]
        if tipo_gd:
            cond_desc.append("tipo_gd = ?")
            params_desc.append(tipo_gd)
        if modal == "AC":
            cond_desc.append("modalidade = 'Autoconsumo'")
        elif modal == "GC":
            cond_desc.append("modalidade = 'Geração Compartilhada'")

        desc_row = conn.execute(
            f"SELECT AVG(desconto_base) AS desconto_gd FROM faturas WHERE {' AND '.join(cond_desc)}",
            params_desc
        ).fetchone()

        registros = conn.execute(f"""
            SELECT
                f.instalacao,
                f.tipo_gd,
                f.modalidade,
                f.desconto_base,
                f.tarifa_compensada,
                f.tarifa_distribuidora
            FROM faturas f
            WHERE {where}
              AND f.tarifa_compensada    IS NOT NULL
              AND f.tarifa_distribuidora IS NOT NULL
            ORDER BY (f.instalacao LIKE 'HIST-%'), f.instalacao
        """, params).fetchall()

    if not row or not row["total"]:
        return jsonify({"erro": "Nenhuma fatura encontrada com os filtros selecionados"}), 404

    modal_short = {"Autoconsumo": "Autocon.", "Geração Compartilhada": "G. Comp."}
    det = [
        {
            "instalacao":          r["instalacao"],
            "tipo_gd":             r["tipo_gd"] or "—",
            "modalidade_short":    modal_short.get(r["modalidade"], r["modalidade"] or "—"),
            "desconto_base":       r["desconto_base"],
            "tarifa_compensada":   round(r["tarifa_compensada"], 6)   if r["tarifa_compensada"]   else None,
            "tarifa_distribuidora": round(r["tarifa_distribuidora"], 6) if r["tarifa_distribuidora"] else None,
        }
        for r in registros
    ]

    # Desconto: 1) clientes reais cadastrados  2) tabela descontos_padrao
    desc_gd = desc_row["desconto_gd"] if desc_row and desc_row["desconto_gd"] is not None else None
    desc_origem = "clientes"
    if desc_gd is None:
        desc_gd = db.get_desconto_padrao(dist)
        desc_origem = "padrao" if desc_gd is not None else None

    return jsonify({
        "tarifa_compensada":    round(row["tarifa_compensada"],    6),
        "tarifa_distribuidora": round(row["tarifa_distribuidora"], 6),
        "tarifa_geracao":       round(row["tarifa_geracao"], 6) if row["tarifa_geracao"] else None,
        "desconto_gd":          round(desc_gd, 6) if desc_gd is not None else None,
        "desconto_origem":      desc_origem,
        "total":                row["total"],
        "total_reais":          row["total_reais"],
        "registros":            det,
    })


# Endpoint temporário para upload do banco de dados (protegido por token)
@app.route("/admin/debug-log")
def admin_debug_log():
    try:
        return DEBUG_LOG.read_text(encoding="utf-8", errors="replace")[-8000:], 200, {"Content-Type": "text/plain; charset=utf-8"}
    except Exception as e:
        return f"Erro: {e}", 500


@app.route("/admin/upload-db", methods=["POST"])
def upload_db():
    token = request.headers.get("X-Admin-Token", "")
    expected = os.environ.get("ADMIN_TOKEN", "")
    if not expected or token != expected:
        return "Unauthorized", 401
    import db as _db
    data = request.data
    if not data:
        return "No data", 400
    with open(_db.DB_PATH, "wb") as f:
        f.write(data)
    return f"OK - {len(data)} bytes gravados em {_db.DB_PATH}", 200


@app.route("/admin/recover-faturas", methods=["POST"])
def admin_recover_faturas():
    token = request.headers.get("X-Admin-Token", "")
    expected = os.environ.get("ADMIN_TOKEN", "")
    if not expected or token != expected:
        return "Unauthorized", 401
    from flask import jsonify
    with db.get_conn() as conn:
        # verifica se faturas_old existe
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if "faturas_old" not in tables:
            count = conn.execute("SELECT COUNT(*) FROM faturas").fetchone()[0]
            return jsonify({"msg": "faturas_old não encontrada", "faturas_count": count})
        # conta registros
        old_count = conn.execute("SELECT COUNT(*) FROM faturas_old").fetchone()[0]
        cur_count = conn.execute("SELECT COUNT(*) FROM faturas").fetchone()[0]
        # copia (INSERT OR IGNORE para não duplicar IDs já presentes)
        old_info = conn.execute("PRAGMA table_info(faturas_old)").fetchall()
        new_info = conn.execute("PRAGMA table_info(faturas)").fetchall()
        old_cols = [c["name"] for c in old_info]
        new_cols = [c["name"] for c in new_info]
        shared = [c for c in old_cols if c in new_cols]
        cols_sql = ", ".join(shared)
        conn.execute(f"INSERT OR IGNORE INTO faturas ({cols_sql}) SELECT {cols_sql} FROM faturas_old")
        recovered = conn.execute("SELECT COUNT(*) FROM faturas").fetchone()[0]
        return jsonify({"ok": True, "recovered": recovered, "faturas_old": old_count, "cols_copied": len(shared)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
